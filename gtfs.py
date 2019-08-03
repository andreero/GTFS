from openpyxl import load_workbook
from collections import namedtuple
from datetime import datetime
import csv
import argparse

AGENCY_ABBREVIATIONS = {
    'Autopromet d.d. Slunj': 'ASLU',
    'Autotrans d.d.': 'AUTT',
    'App d.d. Požega': 'APOE',
    'Panturist d.d. Osijek': 'PANT',
}

ROUTE_TYPE = 3

Agency = namedtuple('Agency', ['agency_id', 'agency_name', 'agency_url', 'agency_timezone'])
Stop = namedtuple('Stop', ['stop_id', 'stop_name', 'stop_desc', 'stop_lat', 'stop_lon', 'stop_timezone'])
Trip = namedtuple('Trip', ['route_id', 'service_id', 'trip_id', 'direction_id', 'block_id', 'shape_id'])
Route = namedtuple('Route', ['route_id', 'agency_id', 'route_long_name', 'route_type'])
Calendar_date = namedtuple('Calendar_date', ['service_id', 'date', 'exception_type'])

agencies = dict()
routes = dict()
trips = dict()


def read_stops_from_file(filename):
    """ Read excel file and return a list of rows """
    workbook = load_workbook(filename)
    active_sheet = workbook.active
    raw_data = []

    for row in range(1, active_sheet.max_row+1):
        cell_value = active_sheet.cell(row, 1).value
        if cell_value is not None:
            raw_data.append(cell_value)
    return raw_data


def process_stops(raw_data):
    reader = csv.reader(raw_data, dialect='excel')
    headers = next(reader, None)
    stop_record = namedtuple("Stop_record", headers)

    stops_dict = dict()
    for row in reader:
        record = stop_record(*row)
        stop = Stop(
            stop_id=record.stop_id,
            stop_name=record.stop_name,
            stop_desc=record.stop_desc,
            stop_lat=record.stop_lat,
            stop_lon=record.stop_lon,
            stop_timezone=record.stop_timezone
        )
        stops_dict[record.place_code] = stop
    return stops_dict


def read_schedule(filename):
    schedule = []
    with open(filename, 'r', encoding='utf-8') as infile:
        reader = csv.reader(infile, dialect='excel', delimiter=';')
        headers = next(reader, None)
        headers[0] = headers[0].replace(u'\ufeff', '')    # remove byte-order-mark, if present

        schedule_tuple = namedtuple("Schedule_record", headers)
        for row in reader:
            record = schedule_tuple(*row)
            schedule.append(record)
    return schedule


def update_agencies(agency_id, record):
    if agency_id not in agencies:
        agency = Agency(agency_id, record.prijevoznik, '', 'Europe/Zagreb')
        agencies[agency_id] = agency


def update_routes(route_id, record):
    if route_id not in routes:
        departure_stop_name = stops.get(record.id_ulaz).stop_name
        arrival_stop_name = stops.get(record.id_izlaz).stop_name
        route = Route(
            route_id=route_id,
            agency_id=AGENCY_ABBREVIATIONS.get(record.prijevoznik, ''),
            route_long_name=f'{departure_stop_name} - {arrival_stop_name}',
            route_type=ROUTE_TYPE
        )
        routes[route_id] = route


def update_trips(record):
    pass


def update_calendar_dates(record):
    pass


def process_schedule(schedule):
    for record in schedule:

        departure = datetime.strptime(record.polazak, '%b  %d %Y  %I:%M%p')
        arrival = datetime.strptime(record.dolazak, '%b  %d %Y  %I:%M%p')
        departure_date = departure.date()
        departure_time = departure.time()
        arrival_time = arrival.time()

        agency_id = AGENCY_ABBREVIATIONS.get(record.prijevoznik, '')
        update_agencies(agency_id, record)

        route_id = '_'.join((agency_id, record.id_ulaz, record.id_izlaz, record.cijena))
        update_routes(route_id, record)


raw_stops = read_stops_from_file('stops.xlsx')
stops = process_stops(raw_stops)

schedule = read_schedule('test.csv')
process_schedule(schedule)
breakpoint()
